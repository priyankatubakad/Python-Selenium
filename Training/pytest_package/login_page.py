import time
from  selenium import  webdriver
import openpyxl

driver=webdriver.Chrome()
driver.get("https://48.ie/log-in")
time.sleep(2)
driver.maximize_window()

path= r'C:\Users\vivek\OneDrive\Desktop\Priyanka\Python_Automation\Python-Selenium\user_details.xlsx'

workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet1"]

email=driver.find_element('css selector', 'input[id="email"]')
pwd=driver.find_element('css selector', 'input[id="pwd-type"]')
signUp=driver.find_element("xpath",
                        '/html/body/app-root/div/div/div[2]/div/app-login/div/form/div[5]/div[1]/button')

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):  # skip header
    print(row)
    email.send_keys(row[0])
    time.sleep(1)
    pwd.send_keys(row[1])
    time.sleep(1)
    signUp.click()
    time.sleep(3)
    email.clear()
    time.sleep(1)
    pwd.clear()
    time.sleep(1)

driver.close()